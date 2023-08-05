import logging
from contextlib import closing
from typing import List

from openpyxl import load_workbook

from submission_broker.submission.entity import Entity
from .clean import clean_entity_name, clean_name, is_value_populated
from .submission import ExcelSubmission

POSSIBLE_KEYS = ['alias', 'index', 'name']


class ExcelLoader:
    def __init__(self, excel_path: str, service_map: dict, service_names: list, sheet_index=0 ):
        # ToDo: Accept param for number of header rows, columns
        self.__path = excel_path
        self.__sheet_index = sheet_index
        self.service_name_map = self.__initial_service_names(service_names)
        self.service_map = service_map
        with closing(load_workbook(filename=self.__path, read_only=True, keep_links=False)) as workbook:
            worksheet = workbook.worksheets[self.__sheet_index]
            self.column_map = self.get_column_map(worksheet)
            self.data = self.get_data(worksheet, self.column_map)

    @staticmethod
    def get_column_map(worksheet) -> dict:
        # Uses iter_rows for faster reads, requires workbook read_only=True
        column_map = {}
        header_rows = []
        object_name = False
        for row in worksheet.iter_rows(min_col=2, max_row=5):
            header_rows.append(row)
        for column_index in range(0, len(header_rows[0])):
            object_cell = header_rows[0][column_index]
            attribute_cell = header_rows[1][column_index]
            units_cell = header_rows[4][column_index]
            column_info = {}

            # Update Object Name otherwise use most recent Object found
            if object_cell.value is not None:
                object_name = clean_entity_name(object_cell.value)
            if object_name:
                column_info['object'] = object_name
            if units_cell.value is not None:
                column_info['units'] = units_cell.value
            if attribute_cell.value is not None:
                column_info['attribute'] = clean_name(attribute_cell.value)
                column_map[attribute_cell.column_letter] = column_info
            else:
                logging.warning(
                    f'No heading found for column: {attribute_cell.column_letter}, values will not be imported.')
        return column_map

    def get_data(self, worksheet, column_map: dict) -> ExcelSubmission:
        data = ExcelSubmission()
        # Import cell values into data object
        # Uses .iter_rows for faster reads, requires workbook read_only=True
        row_index = 6
        for row in worksheet.iter_rows(min_row=row_index, min_col=2):
            row_data = {}
            for cell in row:
                if cell.value is not None \
                        and (cell.is_date or not isinstance(cell.value, str)
                             or is_value_populated(cell.value)):
                    if cell.is_date:
                        value = cell.value.date().isoformat()
                    else:
                        value = str(cell.value).strip()

                    if cell.column_letter in column_map:
                        object_name = column_map[cell.column_letter]['object']
                        attribute_name = column_map[cell.column_letter]['attribute']
                        row_data.setdefault(object_name, {})[attribute_name] = value
            for entity_type, attributes in row_data.items():
                self.map_row_entity(data, row_index, entity_type, attributes)

            row_index = row_index + 1
        return data

    def map_row_entity(self, submission: ExcelSubmission, row: int, entity_type: str, attributes: dict) -> Entity:
        accession_attribute = self.default_accession_attribute(entity_type)
        accession = attributes.get(accession_attribute, None)
        if accession:
            index = accession
        else:
            index = self.get_index(entity_type, row, attributes)
        entity = submission.map_row(row, entity_type, index, attributes)
        if accession and entity_type in self.service_map:
            entity.add_accession(self.service_map[entity_type], accession)
        self.add_entity_accessions(entity, ignore=[accession_attribute])
        return entity

    def get_accession_attribute(self, entity_type: str, service: str):
        if entity_type in self.service_map and service == self.service_map[entity_type]:
            return ExcelLoader.default_accession_attribute(entity_type)

        return ExcelLoader.service_accession_attribute(entity_type, service)

    @staticmethod
    def default_accession_attribute(entity_type: str) -> str:
        return f'{entity_type}_accession'

    @staticmethod
    def service_accession_attribute(entity_type: str, service: str):
        return f'{entity_type.lower()}_{service.lower()}_accession'

    @staticmethod
    def get_index(entity_type: str, row: int, attributes: dict) -> str:
        # Find index in the form 'study_alias', study_index, study_name, ect
        for possible_key in POSSIBLE_KEYS:
            typed_key = f'{entity_type}_{possible_key}'
            if typed_key in attributes:
                return attributes[typed_key]
        # Else: no index found use entity_type:row
        return f'{entity_type}:{row}'

    def add_entity_accessions(self, entity: Entity, ignore: List[str]):
        prefix = f'{entity.identifier.entity_type}_'
        suffix = '_accession'
        attribute: str
        for attribute in entity.attributes.keys():
            if (
                    attribute not in ignore and
                    attribute.startswith(prefix) and
                    attribute.endswith(suffix)
            ):
                service_name = attribute[len(prefix):len(attribute)-len(suffix)]
                if service_name in self.service_name_map:
                    service_name = self.service_name_map[service_name]
                if service_name and entity.attributes[attribute]:
                    entity.add_accession(service_name, entity.attributes[attribute])

    @staticmethod
    def __initial_service_names(service_names: list):
        service_name_map = {}
        for service_name in service_names:
            service_name_map[service_name.lower()] = service_name

        return service_name_map
