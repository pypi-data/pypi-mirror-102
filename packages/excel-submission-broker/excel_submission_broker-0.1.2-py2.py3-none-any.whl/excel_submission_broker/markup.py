from contextlib import closing
from typing import List

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from .validate import ValidatingExcel


class ExcelMarkup(ValidatingExcel):
    def __init__(self, excel_path: str, service_map: dict, service_names: list, sheet_index=0):
        self.__clear_markup(excel_path, sheet_index)
        super().__init__(excel_path, service_map, service_names, sheet_index)
        self.__path = excel_path
        self.__sheet_index = sheet_index
        self.__book = load_workbook(filename=excel_path, keep_links=False)
        self.__sheet = self.__book.worksheets[self.__sheet_index]
        self.attribute_map = self.reverse_column_map(self.column_map)

    def close(self):
        self.__book.save(self.__path)
        self.__book.close()

    def markup_with_errors(self):
        error_fill = PatternFill(fill_type='solid', start_color='FF0000')
        for row_index in self.data.get_all_rows():
            row_error_count = 0
            for entity_type, entity_errors in self.data.get_row_errors(row_index).items():
                row_error_count = row_error_count + len(entity_errors)
                for attribute_name, attribute_errors in entity_errors.items():
                    column_letter = self.get_column_letter(entity_type, attribute_name, 'A')
                    cell_index = self.get_cell_index(column_letter, row_index)
                    self.__sheet[cell_index].comment = \
                        self.__get_error_comment(attribute_errors,self.__sheet[cell_index].comment)
                    self.__sheet[cell_index].fill = error_fill
            if row_error_count:
                self.__sheet[f'A{row_index}'] = f'{row_error_count} Errors'
                self.__sheet[f'A{row_index}'].fill = error_fill

    def add_ena_submission_index(self):
        entity_type = 'submission'
        index_attribute = 'submission_alias'
        for submission_entity in self.data.get_entities(entity_type):
            column_letter = self.get_column_letter(entity_type, index_attribute)
            for row in self.data.get_rows_from_id(submission_entity.identifier):
                cell_index = self.get_cell_index(column_letter, row)
                self.__sheet[cell_index] = submission_entity.identifier.index

    def add_accessions(self):
        for entity_type, entities in self.data.get_all_entities().items():
            for entity in entities:
                for service, accession in entity.get_accessions():
                    self.add_accession(entity_type, entity.identifier.index, service, accession)

    def add_accession(self, entity_type: str, index: str, service: str, accession: str):
        attribute = self.get_accession_attribute(entity_type, service)
        column_letter = self.get_column_letter(entity_type, attribute)
        for row_index in self.data.get_rows(entity_type, index):
            cell_index = self.get_cell_index(column_letter, row_index)
            self.__sheet[cell_index] = accession

    def get_column_letter(self, entity_type, attribute, default_column=None):
        attribute_key = f'{entity_type}.{attribute}'
        letter = self.attribute_map.get(attribute_key, default_column)
        if not letter:
            letter = self.add_column(entity_type, attribute)
        return letter

    def add_column(self, entity_type: str, attribute: str) -> str:
        column_letter = self.__get_next_column_letter()
        self.__sheet[f'{column_letter}1'] = entity_type
        self.__sheet[f'{column_letter}2'] = attribute
        self.column_map[column_letter] = {
            'object': entity_type,
            'attribute': attribute
        }
        self.attribute_map[f'{entity_type}.{attribute}'] = column_letter
        return column_letter

    def __get_next_column_letter(self) -> str:
        return get_column_letter(self.__sheet.max_column + 1)

    @staticmethod
    def get_cell_index(column_letter, row_index):
        return f'{column_letter}{row_index}'

    @staticmethod
    def reverse_column_map(column_map: dict) -> dict:
        attribute_map = {}
        for letter, info in column_map.items():
            attribute_key = f"{info['object']}.{info['attribute']}"
            attribute_map[attribute_key] = letter
        return attribute_map

    @staticmethod
    def __clear_markup(excel_path, sheet_index):
        with closing(load_workbook(filename=excel_path, keep_links=False)) as book:
            sheet = book.worksheets[sheet_index]
            sheet.delete_cols(1, 1)
            sheet.insert_cols(1, 1)

            for row in sheet.iter_rows():
                for cell in row:
                    cell.fill = PatternFill()
                    cell.comment = None
            book.save(excel_path)

    @staticmethod
    def __get_error_comment(errors: List[str], existing_comment: Comment = None):
        stack = []
        if existing_comment and existing_comment.text:
            stack.append(existing_comment.text)
        stack.extend(errors)
        text = '\r\n'.join(stack)
        comment = Comment(text, 'Validation')
        comment.width = 500
        comment.height = 100
        return comment
